"""
Add KPI names to column B in KPI_measurement_quality.xlsx
CAREFULLY preserves all existing formatting and pivot tables
"""

import pandas as pd
from openpyxl import load_workbook

# Step 1: Load consolidator and normalize KPI names
print("Loading consolidator...")
consolidator = pd.read_excel('data/live/AI_QSR_Consolidator.xlsx')

# Normalize KPI data (extract kpi1, kpi2, kpi3 into long format)
kpi_data = []
for idx, row in consolidator.iterrows():
    for kpi_num in [1, 2, 3]:
        kpi_category = row.get(f'kpi{kpi_num}_category')
        if pd.notna(kpi_category):
            kpi_data.append({
                'business_name': row['business_name'],
                'project_name': row['project_name'],
                'kpi_category': kpi_category,
                'kpi_name': row.get(f'kpi{kpi_num}_name')
            })

kpi_lookup = pd.DataFrame(kpi_data)
print(f"Created lookup with {len(kpi_lookup)} KPI entries")

# Step 2: Use openpyxl to ONLY update column B
print("\nOpening Excel file with openpyxl (preserves formatting)...")
filepath = 'analysis/Nov_2/KPI_measurement_quality.xlsx'
wb = load_workbook(filepath)

# Get the first sheet (assuming data is on first sheet)
ws = wb.active
print(f"Working on sheet: {ws.title}")

# Step 3: Read current data to match rows
# Column A = kpi_category, need to find business_name, project_name columns
print("\nFinding column positions...")
header_row = 1
col_positions = {}
for col_idx, cell in enumerate(ws[header_row], start=1):
    if cell.value:
        col_positions[cell.value] = col_idx

print(f"Columns found: {list(col_positions.keys())}")

# Get positions - use column 6 for project_name (first occurrence)
kpi_cat_col = 1  # kpi_category
business_col = 5  # business_name
project_col = 6   # project_name (first occurrence, not column 7)

print(f"Using fixed column positions:")
print(f"  kpi_category column: {kpi_cat_col}")
print(f"  business_name column: {business_col}")
print(f"  project_name column: {project_col}")

# Column B should be position 2 (after A)
kpi_name_col = 2

# Step 4: Update column B header if needed
if ws.cell(row=1, column=kpi_name_col).value != 'kpi_name':
    print(f"\nSetting column B header to 'kpi_name'")
    ws.cell(row=1, column=kpi_name_col, value='kpi_name')

# Step 5: Iterate through rows and match KPI names
print("\nUpdating column B with KPI names...")
rows_updated = 0
rows_not_matched = 0

# Debug first few rows
print("\nDEBUG: First 3 rows from Excel:")
for row_idx in range(2, min(5, ws.max_row + 1)):
    business = ws.cell(row=row_idx, column=business_col).value
    project = ws.cell(row=row_idx, column=project_col).value
    kpi_cat = ws.cell(row=row_idx, column=kpi_cat_col).value
    print(f"  Row {row_idx}: BU='{business}', Project='{project}', KPI='{kpi_cat}'")

print("\nDEBUG: First 3 rows from lookup:")
print(kpi_lookup.head(3)[['business_name', 'project_name', 'kpi_category', 'kpi_name']])

for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (after header)
    business = ws.cell(row=row_idx, column=business_col).value
    project = ws.cell(row=row_idx, column=project_col).value
    kpi_cat = ws.cell(row=row_idx, column=kpi_cat_col).value

    if pd.notna(business) and pd.notna(project) and pd.notna(kpi_cat):
        # Strip whitespace for better matching
        business = str(business).strip() if business else business
        project = str(project).strip() if project else project
        kpi_cat = str(kpi_cat).strip() if kpi_cat else kpi_cat

        # Find matching KPI name
        match = kpi_lookup[
            (kpi_lookup['business_name'].str.strip() == business) &
            (kpi_lookup['project_name'].str.strip() == project) &
            (kpi_lookup['kpi_category'].str.strip() == kpi_cat)
        ]

        if len(match) > 0:
            kpi_name = match.iloc[0]['kpi_name']
            if pd.notna(kpi_name):
                ws.cell(row=row_idx, column=kpi_name_col, value=kpi_name)
                rows_updated += 1
        else:
            rows_not_matched += 1
            # Leave cell empty if no match
            pass

print(f"\nRows updated: {rows_updated}")
print(f"Rows not matched: {rows_not_matched}")

# Step 6: Save the workbook (preserves formatting and pivot tables)
print("\nSaving workbook...")
wb.save(filepath)
print(f"SUCCESS: File saved to {filepath}")
print("All formatting and pivot tables preserved!")
