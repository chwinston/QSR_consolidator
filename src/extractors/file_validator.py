"""
File validator for Excel workbooks.
Implements 6-gate validation to catch errors before extraction.
Note: Merged cells are allowed (businesses submit with merged cells for UX).
"""

from pathlib import Path
from typing import Tuple, List, Optional
from zipfile import ZipFile, BadZipFile
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class ValidationError(Exception):
    """Raised when file validation fails."""

    pass


class FileValidator:
    """
    Validates Excel workbooks before extraction.

    Implements 6 validation gates:
    1. File existence
    2. Extension check (.xlsx or .xlsm)
    3. File size check (reasonable limits)
    4. ZIP integrity (Excel files are ZIP archives)
    5. Password protection check
    6. Sheet structure validation

    Note: Merged cells are now ALLOWED (businesses submit with merged cells for better UX).
    openpyxl can read from merged cells without issues.
    """

    def __init__(
        self,
        min_size_bytes: int = 1024,  # 1 KB
        max_size_bytes: int = 50 * 1024 * 1024,  # 50 MB
        expected_extensions: tuple = ('.xlsx', '.xlsm')
    ):
        """
        Initialize file validator.

        Args:
            min_size_bytes: Minimum acceptable file size
            max_size_bytes: Maximum acceptable file size
            expected_extensions: Tuple of valid file extensions
        """
        self.min_size = min_size_bytes
        self.max_size = max_size_bytes
        self.expected_extensions = expected_extensions

    def validate(self, file_path: Path) -> Tuple[bool, Optional[str]]:
        """
        Run all validation gates.

        Args:
            file_path: Path to Excel file

        Returns:
            Tuple of (is_valid, error_message)
            - (True, None) if validation passes
            - (False, error_message) if validation fails
        """
        try:
            self._validate_file_exists(file_path)
            self._validate_extension(file_path)
            self._validate_file_size(file_path)
            self._validate_zip_integrity(file_path)
            self._validate_not_password_protected(file_path)
            self._validate_sheet_structure(file_path)
            # Note: Merged cells validation disabled - businesses submit with merged cells for UX
            # self._validate_no_merged_cells(file_path)

            return (True, None)

        except ValidationError as e:
            return (False, str(e))

    def _validate_file_exists(self, file_path: Path) -> None:
        """Gate 1: Check file exists."""
        if not file_path.exists():
            raise ValidationError(f"File not found: {file_path}")

        if not file_path.is_file():
            raise ValidationError(f"Path is not a file: {file_path}")

    def _validate_extension(self, file_path: Path) -> None:
        """Gate 2: Check file extension."""
        if not file_path.suffix.lower() in self.expected_extensions:
            raise ValidationError(
                f"Invalid file extension: {file_path.suffix}. "
                f"Expected one of: {', '.join(self.expected_extensions)}"
            )

    def _validate_file_size(self, file_path: Path) -> None:
        """Gate 3: Check file size is within reasonable limits."""
        size = file_path.stat().st_size

        if size < self.min_size:
            raise ValidationError(
                f"File too small ({size} bytes). Minimum: {self.min_size} bytes. "
                "File may be corrupted or empty."
            )

        if size > self.max_size:
            raise ValidationError(
                f"File too large ({size} bytes). Maximum: {self.max_size} bytes."
            )

    def _validate_zip_integrity(self, file_path: Path) -> None:
        """Gate 4: Check ZIP integrity (Excel files are ZIP archives)."""
        try:
            with ZipFile(file_path, 'r') as zip_file:
                # Test ZIP integrity
                bad_file = zip_file.testzip()
                if bad_file:
                    raise ValidationError(
                        f"Corrupted ZIP archive. Bad file: {bad_file}"
                    )
        except BadZipFile:
            raise ValidationError(
                "File is not a valid ZIP archive. Excel file may be corrupted."
            )
        except Exception as e:
            raise ValidationError(f"ZIP validation failed: {str(e)}")

    def _validate_not_password_protected(self, file_path: Path) -> None:
        """Gate 5: Check file is not password protected."""
        try:
            # Try to open with openpyxl
            wb = load_workbook(file_path, read_only=True, data_only=True)
            wb.close()
        except InvalidFileException as e:
            if 'password' in str(e).lower() or 'encrypted' in str(e).lower():
                raise ValidationError(
                    "File is password-protected. Please provide an unprotected file."
                )
            raise ValidationError(f"Cannot open Excel file: {str(e)}")
        except Exception as e:
            raise ValidationError(f"File open failed: {str(e)}")

    def _validate_sheet_structure(self, file_path: Path) -> None:
        """
        Gate 6: Validate sheet structure.

        Checks:
        - At least one project sheet (P1-P10) exists
        - Sheet count is reasonable (1-20 sheets)
        """
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)

            # Check sheet count
            sheet_count = len(wb.sheetnames)
            if sheet_count == 0:
                wb.close()
                raise ValidationError("Workbook contains no sheets.")

            if sheet_count > 20:
                wb.close()
                raise ValidationError(
                    f"Workbook contains too many sheets ({sheet_count}). Maximum: 20."
                )

            # Check for at least one project sheet (P1-P10)
            project_sheets = [name for name in wb.sheetnames if self._is_project_sheet(name)]

            if not project_sheets:
                wb.close()
                raise ValidationError(
                    "No project sheets found. Expected at least one sheet named P1, P2, ..., or P10."
                )

            wb.close()

        except ValidationError:
            raise
        except Exception as e:
            raise ValidationError(f"Sheet structure validation failed: {str(e)}")

    def _validate_no_merged_cells(self, file_path: Path) -> None:
        """
        Gate 7: Check for merged cells in project sheets.

        Merged cells cause extraction issues and should be rejected.
        """
        try:
            wb = load_workbook(file_path, read_only=False, data_only=True)

            # Check only project sheets
            for sheet_name in wb.sheetnames:
                if self._is_project_sheet(sheet_name):
                    ws = wb[sheet_name]

                    if ws.merged_cells:
                        merged_ranges = list(ws.merged_cells.ranges)
                        wb.close()
                        raise ValidationError(
                            f"Sheet '{sheet_name}' contains merged cells: {merged_ranges[0]}. "
                            "Please unmerge all cells before submitting."
                        )

            wb.close()

        except ValidationError:
            raise
        except Exception as e:
            raise ValidationError(f"Merged cells validation failed: {str(e)}")

    @staticmethod
    def _is_project_sheet(sheet_name: str) -> bool:
        """Check if sheet name matches project pattern (P1-P10)."""
        import re
        pattern = r'^P([1-9]|10)$'
        return bool(re.match(pattern, sheet_name))

    def get_project_sheets(self, file_path: Path) -> List[str]:
        """
        Get list of project sheet names in the workbook.

        Args:
            file_path: Path to Excel file

        Returns:
            List of project sheet names (e.g., ['P1', 'P2', 'P3'])

        Raises:
            ValidationError: If file cannot be opened
        """
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            project_sheets = [
                name for name in wb.sheetnames
                if self._is_project_sheet(name)
            ]
            wb.close()
            return project_sheets

        except Exception as e:
            raise ValidationError(f"Failed to get project sheets: {str(e)}")
