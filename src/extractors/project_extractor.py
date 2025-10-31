"""
Project extractor for Excel workbooks.
Extracts 76 fields from project sheets (P1-P10) using config-driven field mapping.
"""

from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.utils.config_loader import ConfigLoader, FieldMapping
from src.utils.error_collector import ErrorCollector
from src.utils.logger import ETLLogger
from src.utils.helpers import (
    parse_date,
    parse_number,
    clean_text,
    generate_extraction_id,
    compute_data_hash,
    get_submission_week
)


class ProjectExtractor:
    """
    Extracts project data from Excel workbooks using field mapping configuration.

    Uses config-driven approach:
    - field_mapping.csv defines which cells to extract
    - Extracts all 76 fields per project sheet
    - Handles data type conversion
    - Collects errors without stopping extraction
    """

    def __init__(
        self,
        config_loader: ConfigLoader,
        logger: Optional[ETLLogger] = None
    ):
        """
        Initialize project extractor.

        Args:
            config_loader: ConfigLoader instance with field mappings loaded
            logger: ETLLogger instance for logging
        """
        self.config = config_loader
        self.logger = logger or ETLLogger()

        if not self.config.field_mappings:
            raise ValueError("ConfigLoader must have field_mappings loaded")

    def extract_workbook(
        self,
        file_path: Path,
        business_id: str,
        sheet_names: Optional[List[str]] = None,
        submission_date: Optional[datetime] = None
    ) -> Tuple[List[Dict[str, Any]], ErrorCollector]:
        """
        Extract all projects from a workbook.

        Args:
            file_path: Path to Excel workbook
            business_id: Business unit identifier
            sheet_names: List of sheet names to extract (default: all P1-P10 sheets)
            submission_date: Submission datetime (default: current datetime)

        Returns:
            Tuple of (extracted_projects, error_collector)
            - extracted_projects: List of project dictionaries
            - error_collector: ErrorCollector with any errors encountered
        """
        if submission_date is None:
            submission_date = datetime.now()

        extraction_id = generate_extraction_id()
        submission_week = get_submission_week(submission_date)

        error_collector = ErrorCollector(partial_success_threshold=0.5)

        self.logger.log_extraction_start(
            extraction_id=extraction_id,
            business_id=business_id,
            file_path=str(file_path)
        )

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)

            # Determine which sheets to extract
            if sheet_names is None:
                sheet_names = [name for name in wb.sheetnames if self._is_project_sheet(name)]

            projects = []

            for sheet_name in sheet_names:
                if sheet_name not in wb.sheetnames:
                    error_collector.add_error(
                        component="ProjectExtractor",
                        error_type="SheetNotFound",
                        message=f"Sheet '{sheet_name}' not found in workbook",
                        context={'file_path': str(file_path), 'sheet_name': sheet_name}
                    )
                    error_collector.record_failure()
                    continue

                try:
                    ws = wb[sheet_name]
                    project_data = self._extract_sheet(
                        ws,
                        sheet_name,
                        business_id,
                        extraction_id,
                        submission_date,
                        submission_week,
                        str(file_path),
                        error_collector
                    )

                    if project_data:
                        projects.append(project_data)
                        error_collector.record_success()
                    else:
                        error_collector.record_failure()

                except Exception as e:
                    error_collector.add_error(
                        component="ProjectExtractor",
                        error_type="SheetExtractionError",
                        message=f"Failed to extract sheet '{sheet_name}': {str(e)}",
                        context={'file_path': str(file_path), 'sheet_name': sheet_name},
                        exception=e
                    )
                    error_collector.record_failure()

            wb.close()

            self.logger.log_extraction_complete(
                extraction_id=extraction_id,
                project_count=len(projects),
                error_count=error_collector.error_count()
            )

            return (projects, error_collector)

        except Exception as e:
            error_collector.add_error(
                component="ProjectExtractor",
                error_type="WorkbookOpenError",
                message=f"Failed to open workbook: {str(e)}",
                context={'file_path': str(file_path)},
                exception=e
            )
            return ([], error_collector)

    def _extract_sheet(
        self,
        ws: Worksheet,
        sheet_name: str,
        business_id: str,
        extraction_id: str,
        submission_date: datetime,
        submission_week: str,
        file_path: str,
        error_collector: ErrorCollector
    ) -> Optional[Dict[str, Any]]:
        """
        Extract all fields from a single project sheet.

        Args:
            ws: Worksheet object
            sheet_name: Name of the sheet
            business_id: Business unit identifier
            extraction_id: Unique extraction ID
            submission_date: Submission datetime
            submission_week: Submission week (YYYY-WW)
            file_path: Path to workbook
            error_collector: ErrorCollector for recording errors

        Returns:
            Dictionary with extracted project data, or None if critical errors
        """
        project_data = {}

        # Add metadata fields
        project_data['extraction_id'] = extraction_id
        project_data['business_id'] = business_id
        project_data['business_name'] = self.config.get_business_by_id(business_id).business_name if self.config.get_business_by_id(business_id) else business_id
        project_data['sheet_name'] = sheet_name
        project_data['submission_date'] = submission_date
        project_data['submission_week'] = submission_week
        project_data['workbook_filename'] = Path(file_path).name

        # Extract all 76 fields
        for field_mapping in self.config.field_mappings:
            field_value = self._extract_field(
                ws,
                field_mapping,
                sheet_name,
                error_collector
            )

            # For KPI numeric fields (target/actual/delta), treat blank cells as 0
            if field_mapping.data_type == 'number' and field_value is None:
                if any(kpi_field in field_mapping.output_column_name
                       for kpi_field in ['kpi1_target', 'kpi1_actual', 'kpi1_delta',
                                        'kpi2_target', 'kpi2_actual', 'kpi2_delta',
                                        'kpi3_target', 'kpi3_actual', 'kpi3_delta']):
                    field_value = 0

            project_data[field_mapping.output_column_name] = field_value

        # Validate required fields
        missing_required = []
        for field_mapping in self.config.get_required_fields():
            if not project_data.get(field_mapping.output_column_name):
                missing_required.append(field_mapping.output_column_name)

        if missing_required:
            error_collector.add_error(
                component="ProjectExtractor",
                error_type="MissingRequiredFields",
                message=f"Required fields missing: {', '.join(missing_required)}",
                context={'sheet_name': sheet_name, 'missing_fields': missing_required}
            )
            return None  # Cannot proceed without required fields

        # Compute data hash for deduplication
        project_name = project_data.get('project_name', '')
        project_data['data_hash'] = compute_data_hash(business_id, project_name, submission_week)

        return project_data

    def _extract_field(
        self,
        ws: Worksheet,
        field_mapping: FieldMapping,
        sheet_name: str,
        error_collector: ErrorCollector
    ) -> Any:
        """
        Extract a single field from worksheet.

        Args:
            ws: Worksheet object
            field_mapping: FieldMapping configuration
            sheet_name: Sheet name (for logging)
            error_collector: ErrorCollector for errors

        Returns:
            Extracted and converted field value
        """
        try:
            # Handle auto-generated/calculated fields (row 0)
            if field_mapping.input_row_number == 0:
                return self._handle_generated_field(field_mapping, ws)

            # Get cell address (e.g., 'C3')
            cell_address = field_mapping.get_cell_address()
            cell_value = ws[cell_address].value

            # Convert based on data type
            if field_mapping.data_type == 'text':
                return clean_text(cell_value)
            elif field_mapping.data_type == 'number':
                return parse_number(cell_value)
            elif field_mapping.data_type == 'date':
                return parse_date(cell_value)
            else:
                return cell_value

        except Exception as e:
            error_collector.add_error(
                component="ProjectExtractor",
                error_type="FieldExtractionError",
                message=f"Failed to extract field '{field_mapping.output_column_name}': {str(e)}",
                context={
                    'sheet_name': sheet_name,
                    'field_name': field_mapping.output_column_name,
                    'cell_address': field_mapping.get_cell_address() if field_mapping.input_row_number > 0 else 'N/A'
                },
                exception=e
            )
            return None

    def _handle_generated_field(
        self,
        field_mapping: FieldMapping,
        ws: Worksheet
    ) -> Any:
        """
        Handle auto-generated or calculated fields (row 0 in field_mapping).

        Args:
            field_mapping: FieldMapping configuration
            ws: Worksheet object

        Returns:
            Generated/calculated value
        """
        import uuid
        from datetime import datetime

        field_name = field_mapping.output_column_name

        # Auto-generate project_uuid
        if field_name == 'project_uuid':
            return str(uuid.uuid4())

        # Calculate total_project_score (will be calculated later in transformer)
        if field_name == 'total_project_score':
            return None  # Will be calculated in data_validator.py

        # additional_context field doesn't exist in template
        if field_name == 'additional_context':
            return None

        # Default: return None for unknown generated fields
        return None

    @staticmethod
    def _is_project_sheet(sheet_name: str) -> bool:
        """Check if sheet name matches project pattern (P1-P10)."""
        import re
        pattern = r'^P([1-9]|10)$'
        return bool(re.match(pattern, sheet_name))
